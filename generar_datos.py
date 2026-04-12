#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generar_datos.py
Genera data.js (CELEBRITIES) y perfume_db.js (PERFUME_DB)
a partir de los xlsx de la base de datos actualizada.
"""

import json
import re
import unicodedata
import openpyxl

def norm_key(s):
    """Normaliza a minúsculas sin tildes para clave de búsqueda."""
    s = s.lower().strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s)
    return s

# ──────────────────────────────────────────────
# 1. PERFUME_DB desde base_datos_perfumes_completa.xlsx
# ──────────────────────────────────────────────
print("Generando perfume_db.js...")

wb_p = openpyxl.load_workbook("base_datos_perfumes_completa.xlsx")
ws_p = wb_p.active

perfume_db = {}
skipped = 0

for r in range(2, ws_p.max_row + 1):
    nombre    = ws_p.cell(r, 1).value
    marca     = ws_p.cell(r, 2).value
    genero    = ws_p.cell(r, 3).value
    estacion  = ws_p.cell(r, 4).value
    momento   = ws_p.cell(r, 5).value
    perfil    = ws_p.cell(r, 6).value
    ocasion   = ws_p.cell(r, 7).value
    familia   = ws_p.cell(r, 8).value
    notas     = ws_p.cell(r, 9).value
    proyeccion= ws_p.cell(r, 10).value
    precio    = ws_p.cell(r, 11).value
    vibe      = ws_p.cell(r, 12).value
    nivel     = ws_p.cell(r, 13).value

    if not nombre:
        continue

    # Normalizar precio para consistencia
    if precio:
        precio = precio.replace("€", "€").strip()

    key = norm_key(str(nombre))
    perfume_db[key] = {
        "nombre":     str(nombre).strip(),
        "marca":      str(marca).strip() if marca else "",
        "genero":     str(genero).strip() if genero else "",
        "estacion":   str(estacion).strip() if estacion else "",
        "momento":    str(momento).strip() if momento else "",
        "perfil":     str(perfil).strip() if perfil else "",
        "ocasion":    str(ocasion).strip() if ocasion else "",
        "familia":    str(familia).strip() if familia else "",
        "notas":      str(notas).strip() if notas else "",
        "proyeccion": str(proyeccion).strip() if proyeccion else "",
        "precio":     str(precio).strip() if precio else "",
        "vibe":       str(vibe).strip() if vibe else "",
        "nivel":      str(nivel).strip() if nivel else "",
    }

print(f"  {len(perfume_db)} perfumes procesados")

with open("perfume_db.js", "w", encoding="utf-8") as f:
    f.write("const PERFUME_DB = ")
    f.write(json.dumps(perfume_db, ensure_ascii=False, separators=(",", ":")))
    f.write(";")

print("  perfume_db.js generado OK")

# ──────────────────────────────────────────────
# 2. CELEBRITIES desde celebrities_perfumes_corregido.xlsx
# ──────────────────────────────────────────────
print("\nGenerando data.js...")

wb_c = openpyxl.load_workbook("celebrities_perfumes_corregido.xlsx")
ws_c = wb_c.active

# Agrupar por celebrity (manteniendo orden de aparición)
celebrities_map = {}
order = []

for r in range(2, ws_c.max_row + 1):
    nombre   = ws_c.cell(r, 1).value
    profesion= ws_c.cell(r, 2).value
    categoria= ws_c.cell(r, 3).value
    perfume  = ws_c.cell(r, 4).value
    marca    = ws_c.cell(r, 5).value

    if not nombre or not perfume:
        continue

    nombre = str(nombre).strip()
    profesion = str(profesion).strip() if profesion else ""
    categoria = str(categoria).strip() if categoria else ""
    perfume = str(perfume).strip()
    marca = str(marca).strip() if marca else ""

    if nombre not in celebrities_map:
        celebrities_map[nombre] = {
            "nombre": nombre,
            "profesion": profesion,
            "categoria": categoria,
            "perfumes": []
        }
        order.append(nombre)

    celebrities_map[nombre]["perfumes"].append({
        "perfume": perfume,
        "marca": marca
    })

celebrities_list = [celebrities_map[n] for n in order]
print(f"  {len(celebrities_list)} celebrities procesados")

with open("data.js", "w", encoding="utf-8") as f:
    f.write("const CELEBRITIES = ")
    f.write(json.dumps(celebrities_list, ensure_ascii=False, separators=(",", ":")))
    f.write(";")

print("  data.js generado OK")
print("\nListo.")
